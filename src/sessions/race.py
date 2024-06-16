import statistics

import openpyxl
from matplotlib.ticker import FuncFormatter
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import matplotlib.patheffects as path_effects
import statsmodels.api as sm
from statsmodels.robust.norms import HuberT
from statsmodels.robust.robust_linear_model import RLM

from src.ergast_api.my_ergast import My_Ergast
from src.exceptions import race_same_team_exceptions
from src.exceptions.custom_exceptions import RaceException
from src.plots.plots import get_font_properties, round_bars, annotate_bars
import fastf1
from fastf1.ergast import Ergast
from fastf1 import plotting

from matplotlib import pyplot as plt, ticker, patheffects
import seaborn as sns
import numpy as np
import pandas as pd
import matplotlib.colors as mcolors
import matplotlib.patches as mpatches

from src.utils.utils import append_duplicate_number, split_at_discontinuities, format_timedelta
from src.variables.driver_colors import driver_colors_2023, driver_colors
from src.variables.team_colors import team_colors_2023, team_colors
import matplotlib.dates as mdates
from scipy import stats


def position_changes(session):
    """
         Plots the position changes in every lap

         Parameters:
         session (Session): Session to analyze

    """

    plotting.setup_mpl(misc_mpl_mods=False)

    fig, ax = plt.subplots(figsize=(8.0, 5.12))
    for drv in session.drivers:
        if drv != '55':
            drv_laps = session.laps.pick_driver(drv)
            abb = drv_laps['Driver'].iloc[0]
            if abb == 'MSC':
                color = '#cacaca'
            elif abb == 'VET':
                color = '#006f62'
            elif abb == 'LAT':
                color = '#012564'
            elif abb == 'LAW':
                color = '#2b4562'
            else:
                color = plotting.driver_color(abb)

            starting_grid = session.results.GridPosition.to_frame().reset_index(drop=False).rename(
                columns={'index': 'driver'})
            dict_drivers = session.results.Abbreviation.to_dict()

            starting_grid = starting_grid.replace({'driver': dict_drivers}).replace({'GridPosition': {0: 20}})

            vueltas = drv_laps['LapNumber'].reset_index(drop=True)
            position = drv_laps['Position'].reset_index(drop=True)
            position.at[0] = starting_grid[starting_grid['driver'] == abb]['GridPosition']

            ax.plot(vueltas, position,
                    label=abb, color=color)
    ax.set_xlim([1, session.total_laps])
    ax.set_ylim([20.5, 0.5])
    ax.set_yticks([1, 5, 10, 15, 20])
    ax.set_xlabel('Lap')
    ax.set_ylabel('Position')
    ax.legend(bbox_to_anchor=(1.0, 1.02))

    plt.title(session.event.OfficialEventName + ' ' + session.name.upper(), font='Fira Sans', fontsize=18, loc='center')
    plt.figtext(0.01, 0.02, '@Big_Data_Master', fontsize=15, color='gray', alpha=0.5)
    plt.tight_layout()
    plt.savefig(f"../PNGs/POSITION CHANGES {session.event.OfficialEventName}.png", dpi=400)
    plt.show()


def driver_race_times_per_tyre(session, driver):
    """
         Plots all the time laps, with the compound, for a driver

         Parameters:
         race (Session): Session to analyze
         driver (str): Driver

    """

    fastf1.plotting.setup_mpl(misc_mpl_mods=False)
    driver_laps = session.laps.pick_driver(driver).pick_quicklaps().reset_index()
    driver_laps["LapTime(s)"] = driver_laps["LapTime"].dt.total_seconds()

    fig, ax = plt.subplots(figsize=(8, 8))

    sns.scatterplot(data=driver_laps,
                    x="LapNumber",
                    y="LapTime(s)",
                    ax=ax,
                    hue="Compound",
                    palette=fastf1.plotting.COMPOUND_COLORS,
                    s=80,
                    linewidth=0,
                    legend='auto')

    ax.set_xlabel("Lap Number", font='Fira Sans', fontsize=16)
    ax.set_ylabel("Lap Time", font='Fira Sans', fontsize=16)

    def lap_time_formatter(x, pos):
        minutes, seconds = divmod(x, 60)
        return f"{int(minutes):02}:{seconds:06.3f}"

    ax.yaxis.set_major_formatter(ticker.FuncFormatter(lap_time_formatter))
    ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))

    plt.suptitle(f"{driver} Laptimes in {session.event.OfficialEventName}", font='Fira Sans', fontsize=14)

    # Turn on major grid lines
    plt.grid(color='w', which='major', axis='both')
    sns.despine(left=True, bottom=True)

    plt.tight_layout()
    plt.savefig(f"../PNGs/RACE LAPS {driver} {session.event.OfficialEventName}.png", dpi=500)
    plt.show()


def tyre_strategies(session):
    """
         Plots the tyre strategy in a race

         Parameters:
         session (Session): Session to analyze

    """

    laps = session.laps
    drivers = session.drivers

    drivers = [session.get_driver(driver)["Abbreviation"] for driver in drivers]
    stints = laps[["Driver", "Stint", "Compound", "LapNumber", "FreshTyre", "TyreLife"]]
    past_stint = 1
    past_driver = stints.loc[0, 'Driver']
    for i in range(len(stints)):
        changed = False
        current_driver = stints.loc[i, 'Driver']

        if past_driver != current_driver:
            past_stint = 1

        current_stint = stints.loc[i, 'Stint']
        if past_stint != current_stint and i != 0:
            tyre_laps_prev = stints.loc[i - 1, 'TyreLife']
            tyre_laps = stints.loc[i, 'TyreLife']
            compound_prev = stints.loc[i - 1, 'Compound']
            compound = stints.loc[i, 'Compound']

            if tyre_laps_prev + 1 == tyre_laps and (compound_prev == compound):
                indexes = stints.index[stints['Driver'] == stints.loc[i, 'Driver']].tolist()
                indexes = [x for x in indexes if x >= i]
                for index in indexes:
                    stints.loc[index, 'Stint'] = stints.loc[index, 'Stint'] - 1
                    stints.loc[index, 'FreshTyre'] = stints.loc[min(indexes) - 1, 'FreshTyre']
                past_stint = current_stint - 1
                changed = True

        if not changed:
            past_stint = current_stint
            past_driver = current_driver

    stints = stints.groupby(["Driver", "Stint", "Compound", "FreshTyre"])
    stints = stints.count().reset_index()

    stints = stints.rename(columns={"LapNumber": "StintLength"})

    fig, ax = plt.subplots(figsize=(8, 8))

    patches = {}

    for driver in drivers:
        driver_stints = stints.loc[stints["Driver"] == driver]
        previous_stint_end = 0
        for idx, row in driver_stints.iterrows():
            if row['FreshTyre']:
                alpha = 1
                color = plotting.COMPOUND_COLORS[row["Compound"]]
            else:
                alpha = 0.65
                rgb_color = mcolors.to_rgb(plotting.COMPOUND_COLORS[row["Compound"]])
                color = tuple([x * 0.95 for x in rgb_color])
            stint_text = str(row["StintLength"])
            text_x_position = previous_stint_end + row["StintLength"] / 2  # center the text in the stint bar
            text_y_position = driver  # align vertically with the driver's bar
            text = ax.text(
                text_x_position, text_y_position, stint_text,
                ha='center', va='center', font='Fira Sans', fontsize=13, color='black', alpha=1
            )

            text.set_path_effects([
                patheffects.withStroke(linewidth=2.5, foreground="white")
            ])

            plt.barh(
                y=driver,
                width=row["StintLength"],
                left=previous_stint_end,
                color=color,
                edgecolor="black",
                fill=True,
                alpha=alpha
            )

            label = f"{'New' if row['FreshTyre'] else 'Used'} {row['Compound']}"
            color_rgb = mcolors.to_rgb(color)  # Convert color name to RGB
            patches[label] = mpatches.Patch(color=color_rgb + (alpha,), label=label)
            previous_stint_end += row["StintLength"]

    patches = {k: patches[k] for k in sorted(patches.keys(), key=lambda x: (x.split()[1], x.split()[0]))}
    plt.legend(handles=patches.values(), bbox_to_anchor=(0.5, -0.1),
               prop=get_font_properties('Fira Sans', 'large'), loc='upper center', ncol=2)

    plt.title(f'TYRE STRATEGIES IN {session.event.Location.upper()}', font='Fira Sans', fontsize=26)
    plt.xlabel("Lap Number", font='Fira Sans', fontsize=14)
    plt.grid(False)
    ax.invert_yaxis()

    plt.xticks(font='Fira Sans', fontsize=14)
    plt.yticks(font='Fira Sans', fontsize=14)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)

    plt.tight_layout()
    plt.savefig(f"../PNGs/TYRE STRATEGY {session.event.OfficialEventName}.png", dpi=400)
    plt.show()


def race_pace_top_10(session, threshold=1.07):
    """
    Plots the race pace of the top 10 drivers in a race

    Parameters:
    race (session): Race to be plotted

    """
    fastf1.plotting.setup_mpl(mpl_timedelta_support=False, misc_mpl_mods=False)
    session_year = session.event.year
    drivers = set(session.laps['Driver'])
    total_laps = pd.DataFrame()
    for d in drivers:
        if d not in ['N']:
            d_laps = pd.DataFrame(session.laps.pick_driver(d).pick_quicklaps(threshold).pick_wo_box())
            total_laps = total_laps._append(d_laps)
    total_laps = total_laps.groupby(['Driver', 'Team'])['LapTime'].mean().reset_index()
    total_laps['FastestLapTime'] = total_laps.groupby('Team')['LapTime'].transform(min)
    total_laps['IsFastest'] = total_laps['LapTime'] == total_laps['FastestLapTime']
    total_laps = total_laps[total_laps['IsFastest']]
    total_laps = total_laps.sort_values('LapTime')
    finishing_order = total_laps['Driver'].values

    driver_laps = pd.DataFrame()
    for d in finishing_order:
        d_laps = session.laps.pick_drivers(d).pick_quicklaps(threshold).pick_wo_box()
        driver_laps = driver_laps._append(d_laps)

    team_colors_dict = team_colors.get(session_year)
    driver_colors = []
    for d in finishing_order:
        color = team_colors_dict[session.laps.pick_driver(d)['Team'].loc[0]]
        if color == '#ffffff':
            color = '#FF7C7C'
        driver_colors.append(color)
    fig, ax = plt.subplots(figsize=(8, 6))
    driver_laps["LapTime(s)"] = driver_laps["LapTime"].dt.total_seconds()
    mean_lap_times = driver_laps.groupby("Driver")["LapTime"].mean()

    sns.violinplot(data=driver_laps,
                   x="Driver",
                   y="LapTime(s)",
                   inner=None,
                   scale="area",
                   order=finishing_order,
                   palette=driver_colors,
                   )

    sns.swarmplot(data=driver_laps,
                  x="Driver",
                  y="LapTime(s)",
                  order=finishing_order,
                  hue="Compound",
                  palette=fastf1.plotting.COMPOUND_COLORS,
                  hue_order=["SOFT", "MEDIUM", "HARD"],
                  linewidth=0,
                  size=5,
                  )

    # all_drivers = set(session.laps['Driver'].values)
    # for d in all_drivers:
    #     d_laps = session.laps.pick_driver(d).pick_wo_box().pick_quicklaps(threshold)
    #     print(f'{d} - {format_timedelta(d_laps["LapTime"].mean())}')

    fastest_time = None
    position = 1
    percentage_dict = {}
    for i, driver in enumerate(finishing_order):
        team = session.laps.pick_driver(driver)['Team'].loc[0]
        mean_time = mean_lap_times[driver]
        f_mean_time = format_timedelta(mean_time)
        ax.text(i, min(driver_laps['LapTime(s)']) - 1.5,
                f"{f_mean_time}",
                font='Fira Sans', fontsize=10, ha='center')
        if fastest_time is None:
            time_diff = ''
            percentage_diff = 0.0  # No percentage difference for the fastest driver
            fastest_time = mean_time
        else:
            time_diff_seconds = (mean_time - fastest_time).total_seconds()
            time_diff = f'(+{format_timedelta(mean_time - fastest_time)}s)'.replace('0:0', '')
            percentage_diff = time_diff_seconds / fastest_time.total_seconds() * 100
        percentage_dict[team] = percentage_diff
        print(f'{position} - {team}: {f_mean_time} {time_diff}')
        position += 1

    for t, p in percentage_dict.items():
        print(f'{t},{p}')

    plt.ylim(min(driver_laps['LapTime']).total_seconds() - 2,
             max(driver_laps['LapTime']).total_seconds() + 2)

    def format_timedelta_to_mins(seconds, pos):
        mins = int(seconds // 60)
        secs = int(seconds % 60)
        millis = int((seconds - int(seconds)) * 1000)
        return f'{mins:01d}:{secs:02d}.{millis:03d}'

    formatter = FuncFormatter(format_timedelta_to_mins)
    plt.gca().yaxis.set_major_formatter(formatter)

    plt.xticks(fontsize=15)
    plt.yticks(fontsize=15)
    ax.set_xlabel("Driver", font='Fira Sans', fontsize=17)
    ax.set_ylabel("Lap Time (s)", font='Fira Sans', fontsize=17)
    plt.suptitle(f"{session.event['EventDate'].year} {session.event['EventName']} Lap Time Distributions",
                 font='Fira Sans', fontsize=20)
    sns.despine(left=False, bottom=False)
    plt.legend(title='Tyre Compound', loc='upper left', fontsize='medium')

    plt.tight_layout()
    plt.savefig(f"../PNGs/{session.event['EventDate'].year} {session.event['EventName']} Lap Time Distributions.png",
                dpi=450)
    plt.show()


def race_diff(year):
    """
         Plots the race time diff between 2 different teams

         Parameters:
         year (int): Year to analyze

    """

    session_names = []
    n_races = Ergast().get_race_results(season=year, limit=1000)
    mean_delta_by_team = {}
    for i in range(len(n_races.content)):
        if not (year == 2021 and (i + 1) == 12):
            race = fastf1.get_session(year, i + 1, 'R')
            race.load(telemetry=True)
            total_laps = race.total_laps
            session_names.append(race.event['Location'].split('-')[0])
            teams = race.results['TeamName'].unique()
            teams = teams[teams != 'nan']
            pace_with_laps = {}
            for t in teams:
                drivers = race.laps[race.laps['Team'] == t]['Driver'].unique()
                n_laps = [len(race.laps[race.laps['Driver'] == i]) for i in drivers]
                paces = {}
                for d in drivers:
                    for n in set(n_laps):
                        if n >= total_laps / 3:
                            if len(race.laps.pick_driver(d)[0:n]) >= n:
                                mean = race.laps.pick_driver(d)[0:n].pick_quicklaps().pick_wo_box()['LapTime'].mean()
                                if d not in paces:
                                    paces[d] = {'MEAN': [mean], 'LAPS': [n]}
                                else:
                                    paces[d]['MEAN'].append(mean)
                                    paces[d]['LAPS'].append(n)
                if len(paces) == 0:
                    if t not in mean_delta_by_team:
                        mean_delta_by_team[t] = [np.NaN]
                    else:
                        mean_delta_by_team[t].append(np.NaN)
                else:
                    if len(paces) > 1:
                        paces = dict(sorted(paces.items(),
                                            key=lambda item: (min(item[1]["LAPS"]),
                                                              item[1]["MEAN"][
                                                                  item[1]["LAPS"].index(min(item[1]["LAPS"]))])))
                    else:
                        key = next(iter(paces))
                        laps, means = paces[key]["LAPS"], paces[key]["MEAN"]
                        sorted_pairs = sorted(zip(laps, means), key=lambda pair: pair[1])
                        sorted_laps, sorted_means = zip(*sorted_pairs)
                        paces = {key: {"LAPS": sorted_laps, "MEAN": sorted_means}}
                    pace_with_laps[t] = {"MEAN": paces[list(paces.keys())[0]]['MEAN'][0],
                                         "LAPS": paces[list(paces.keys())[0]]['LAPS'][0]}
            pace_with_laps = dict(sorted(pace_with_laps.items(), key=lambda item: (item[1]["MEAN"])))
            current_fastest_team = list(pace_with_laps.keys())[0]
            current_fastest_mean = pace_with_laps[list(pace_with_laps.keys())[0]]['MEAN']
            pace_with_laps = dict(sorted(pace_with_laps.items(), key=lambda item: (-item[1]["LAPS"], item[1]["MEAN"])))
            for t, m in pace_with_laps.items():
                if m['LAPS'] < total_laps - 5:
                    laps = m['LAPS']
                    pace = m['MEAN']
                    drivers_fastest_team = race.laps[race.laps['Team'] == current_fastest_team]['Driver'].unique()
                    mean_fixed_laps = []
                    for d in drivers_fastest_team:
                        mean_fixed_laps.append(
                            race.laps.pick_driver(d)[0:laps].pick_quicklaps().pick_wo_box()['LapTime'].mean())
                    fastest_fixed_mean = min(mean_fixed_laps)
                    delta_diff = ((pace - fastest_fixed_mean) / fastest_fixed_mean) * 100
                    if t not in mean_delta_by_team:
                        mean_delta_by_team[t] = [delta_diff]
                    else:
                        mean_delta_by_team[t].append(delta_diff)
                elif t != current_fastest_team:
                    team_mean = m['MEAN']
                    delta_diff = ((team_mean - current_fastest_mean) / current_fastest_mean) * 100
                    if t not in mean_delta_by_team:
                        mean_delta_by_team[t] = [delta_diff]
                    else:
                        mean_delta_by_team[t].append(delta_diff)
                elif t == current_fastest_team:
                    if t not in mean_delta_by_team:
                        mean_delta_by_team[t] = [0]
                    else:
                        mean_delta_by_team[t].append(0)
            min_latest_value = min(value[-1] if isinstance(value, list) and len(value) > 0 else 0
                                   for value in mean_delta_by_team.values())
            if min_latest_value != 0:
                for team, value in mean_delta_by_team.items():
                    if len(value) > 0:
                        mean_delta_by_team[team][-1] = value[-1] - min_latest_value
                print(f'Recompute {race.event.Location}')
        else:
            for t, d in mean_delta_by_team.copy().items():
                mean_delta_by_team[t].append(np.NaN)
            session_names.append('Spa')

    for t, d in mean_delta_by_team.items():
        print(f'{t} - {d}')

    session_names = append_duplicate_number(session_names)

    df = pd.DataFrame(mean_delta_by_team)
    df['Year'] = year
    df['Session'] = df.index + 1
    df['Track'] = session_names
    df = df.melt(id_vars=['Year', 'Session', 'Track'], var_name='Team', value_name='Delta')
    existing_data = pd.read_csv('../resources/csv/Raw_race_pace_delta.csv')
    composite_pk_columns = ['Year', 'Session', 'Team']
    merged = pd.merge(existing_data[composite_pk_columns], df, how='outer', on=composite_pk_columns, indicator=True)
    unique_rows = merged[merged['_merge'] == 'right_only'].drop(columns=['_merge']).reset_index(drop=True)
    unique_rows['ID'] = unique_rows.index + max(existing_data['ID']) + 1
    unique_rows = unique_rows[existing_data.columns]
    print(f'{len(unique_rows)} NEW ROWS')
    unique_rows.to_csv('../resources/csv/Raw_race_pace_delta.csv', mode='a', header=False, index=False)

    fig, ax1 = plt.subplots(figsize=(12, 8))
    plt.rcParams["font.family"] = "Fira Sans"
    for team, deltas in mean_delta_by_team.items():
        deltas_array = np.array(deltas)
        not_nan_indices = ~np.isnan(deltas_array)
        plt.plot(np.array(session_names)[not_nan_indices], deltas_array[not_nan_indices],
                 label=team, marker='o', color=team_colors_2023.get(team), markersize=7, linewidth=3)
    plt.gca().invert_yaxis()
    plt.legend(loc='lower right', fontsize='medium')
    plt.title(f'{year} AVERAGE RACE PACE PER CIRCUIT', font='Fira Sans', fontsize=20)
    plt.ylabel('Percentage time difference (%)', font='Fira Sans', fontsize=16)
    plt.xlabel('Circuit', font='Fira Sans', fontsize=16)
    ax1.yaxis.grid(True, linestyle='--')
    ax1.xaxis.grid(True, linestyle='--', alpha=0.2)
    plt.xticks(rotation=90, fontsize=12, fontname='Fira Sans')
    plt.yticks(fontsize=12, fontname='Fira Sans')
    plt.tight_layout()
    plt.savefig(f"../PNGs/{year} race time difference.png", dpi=400)
    plt.show()


def race_distance(session, top=10):
    """
         Plots the race time diff between 2 drivers, each lap

         Parameters:
         session (Session): Session to analyze
         driver_1 (str): Driver 1
         driver_2 (str): Driver 2
    """

    drivers = session.results['Abbreviation'].values[:top]
    laps = session.total_laps
    diff = [[] for _ in range(len(drivers))]
    leader_laps = session.laps.pick_driver(drivers[0])
    leader_laps = leader_laps['Time'].dt.total_seconds().reset_index(drop=True)
    bar_colors = [driver_colors_2023[d] for d in drivers]
    laps_array = [i for i in range(1, laps + 1)]
    sc_laps = session.laps.pick_driver(drivers[0])['TrackStatus']
    sc_laps = sc_laps.astype(str).str.contains('4|5|6|7')
    sc_laps = sc_laps[sc_laps == True].index.values
    sc_laps = session.laps.pick_driver(drivers[0])['LapNumber'][sc_laps].values
    for i in range(len(drivers)):
        diff[i] = []
        d_laps = session.laps.pick_driver(drivers[i])
        d_laps = d_laps['Time'].dt.total_seconds().reset_index(drop=True)
        if i == 0:
            diff[i].extend([0] * laps)
        else:
            delta_diff = d_laps - leader_laps
            diff[i].extend(delta_diff.values)
    fig, ax = plt.subplots(figsize=(8, 8))
    for i in range(len(diff)):
        plt.plot(laps_array, diff[i], color=bar_colors[i], label=drivers[i], linewidth=2.5)
    if len(sc_laps) > 0:
        sc_intervals = split_at_discontinuities(sc_laps)
        for interval in sc_intervals:
            ax.axvspan(interval[0] - 1, interval[-1] - 1, color='yellow', alpha=0.3)

    plt.gca().invert_yaxis()
    plt.xlabel('Laps', font='Fira Sans', fontsize=16)
    plt.ylabel('Progressive Time Difference (seconds)', font='Fira Sans', fontsize=16)
    plt.xticks(font='Fira Sans', fontsize=14)
    plt.yticks(font='Fira Sans', fontsize=14)
    plt.figtext(0.01, 0.02, '@Big_Data_Master', font='Fira Sans', fontsize=15, color='gray', alpha=0.5)
    plt.title(
        f'Progressive Time Difference between in {session.event["EventName"] + " " + str(session.event["EventDate"].year)}',
        font='Fira Sans', fontsize=17)
    plt.grid(True, which='both', linestyle='--', linewidth=0.5)
    plt.legend(prop=get_font_properties('Fira Sans', 12), loc='lower left', fontsize='x-large')
    plt.tight_layout()  # Adjusts plot parameters for a better layout
    plt.savefig(
        f'../PNGs/Progressive Time Difference between in {session.event["EventName"] + " " + str(session.event["EventDate"].year)}',
        dpi=400)
    plt.show()


def race_pace_between_drivers(year, d1, d2, round_id=None, all_teams=False, session_type='R', number_rounds=None):
    schedule = fastf1.get_event_schedule(year, include_testing=False)
    fuel_factors = pd.read_csv('../resources/csv/Fuel_factor.csv')
    schedule = [1] if round_id is not None else schedule
    schedule = schedule if number_rounds is None else [i for i in range(number_rounds)]
    total_laps_d1 = []
    total_laps_d2 = []
    all_comparisons = {}
    sessions_names = []
    for i in range(len(schedule)):
        session = fastf1.get_session(year, i + 1 if round_id is None else round_id, session_type)
        session.load()
        fuel_data_session = i + 1 if round_id is None else round_id
        race_fuel_factor = fuel_factors[(fuel_factors['Round'] == fuel_data_session) &
                                        (fuel_factors['Year'] == year)]['FuelFactor'][0]
        if pd.isna(race_fuel_factor):
            race_fuel_factor = 0.06
        print(f'FUEL FACTOR: {race_fuel_factor}')
        loops = 1
        if all_teams:
            teams = session.laps['Team'].unique()
            teams = [t for t in teams if not pd.isna(t)]
            loops = len(teams)
        for n_loops in range(loops):
            try:
                if all_teams:
                    team = teams[n_loops]
                    drivers = session.laps[session.laps['Team'] == team]['Driver'].unique()
                    if len(drivers) != 2:
                        raise RaceException
                    d1 = drivers[0]
                    d2 = drivers[1]
                team_d1 = session.laps.pick_driver(d1)['Team'].unique()[0]
                team_d2 = session.laps.pick_driver(d2)['Team'].unique()[0]
                from src.utils.utils import call_function_from_module
                try:
                    min_laps_d1, max_laps_d1 = call_function_from_module(race_same_team_exceptions,
                                                                         f"{team_d1.replace(' ', '_')}_{year}",
                                                                         i + 1 if round_id is None else round_id, session.total_laps)
                    min_laps_d2, max_laps_d2 = call_function_from_module(race_same_team_exceptions,
                                                                         f"{team_d2.replace(' ', '_')}_{year}",
                                                                         i + 1 if round_id is None else round_id, session.total_laps)
                except AttributeError:
                    min_laps_d1 = 0
                    max_laps_d1 = session.total_laps
                    min_laps_d2 = 0
                    max_laps_d2 = session.total_laps
                min_laps = max(min_laps_d1, min_laps_d2, 1)
                max_laps = min(max_laps_d1, max_laps_d2, len(session.laps.pick_driver(d1)), len(session.laps.pick_driver(d2)))
                d1_laps = session.laps.pick_driver(d1)[min_laps:max_laps].pick_quicklaps().pick_wo_box()
                d2_laps = session.laps.pick_driver(d2)[min_laps:max_laps].pick_quicklaps().pick_wo_box()
                d1_compare = pd.DataFrame(d1_laps[['Driver', 'LapNumber', 'LapTime', 'Compound', 'TyreLife']])
                d2_compare = pd.DataFrame(d2_laps[['LapNumber', 'LapTime', 'Compound', 'TyreLife', 'Driver']])
                d1_compare['EvoCorrected'] = (d1_compare['LapTime'].dt.total_seconds() - race_fuel_factor *
                                               (session.total_laps - d1_compare['LapNumber']))
                d2_compare['EvoCorrected'] = (d2_compare['LapTime'].dt.total_seconds() - race_fuel_factor *
                                               (session.total_laps - d2_compare['LapNumber']))
                d1_compare.columns = ['Driver 1', 'LapNumber D1', 'LapTime D1', 'Compound', 'TyreLife', 'EvoCorrected D1']
                d2_compare.columns = ['LapNumber D2', 'LapTime D2', 'Compound', 'TyreLife', 'Driver 2', 'EvoCorrected D2']
                comparable_laps = pd.merge(d1_compare, d2_compare, how='inner', on=['Compound', 'TyreLife'])
                total_laps_d1.extend(pd.to_timedelta(comparable_laps['EvoCorrected D1'], unit='s'))
                total_laps_d2.extend(pd.to_timedelta(comparable_laps['EvoCorrected D2'], unit='s'))
                if len(comparable_laps) > 0:
                    sessions_names.append(session.event.Location)
                    sorted_laps_d1 = comparable_laps.sort_values(by='EvoCorrected D1')['EvoCorrected D1'].drop_duplicates()
                    sorted_laps_d2 = comparable_laps.sort_values(by='EvoCorrected D2')['EvoCorrected D2'].drop_duplicates()
                    trim_percent = 0.05
                    num_rows = len(comparable_laps)
                    rows_to_trim = int(trim_percent * num_rows)
                    start_trimmed_laps_d1 = sorted_laps_d1[:rows_to_trim]
                    end_trimmed_laps_d1 = sorted_laps_d1[-rows_to_trim:] if rows_to_trim != 0 else sorted_laps_d1[0:0]
                    trimmed_laps_d1 = pd.concat([start_trimmed_laps_d1, end_trimmed_laps_d1])

                    start_trimmed_laps_d2 = sorted_laps_d2[:rows_to_trim]
                    end_trimmed_laps_d2 = sorted_laps_d2[-rows_to_trim:] if rows_to_trim != 0 else sorted_laps_d2[0:0]
                    trimmed_laps_d2 = pd.concat([start_trimmed_laps_d2, end_trimmed_laps_d2])

                    excluded_d1 = comparable_laps[comparable_laps['EvoCorrected D1'].isin(trimmed_laps_d1)]
                    excluded_d2 = comparable_laps[comparable_laps['EvoCorrected D2'].isin(trimmed_laps_d2)]

                    comparable_laps['Exclude'] = 'No'
                    comparable_laps.loc[comparable_laps.index.isin(excluded_d1.index) | comparable_laps.index.isin(
                        excluded_d2.index), 'Exclude'] = 'Yes'

                    all_comparisons[(f'{team_d1 if round_id is not None else ""} '
                                     f' {" - " + session.event.EventName if not all_teams else ""}')] = comparable_laps
                else:
                    print(f'NO DATA FOR {team_d1}')
            except (RaceException, IndexError):
                print(f'{session}')

    path_to_save = f'../resources/race_pace/{year}/'
    file_name = f"Round - {round_id}" if all_teams else f"{d1} VS. {d2}"
    full_path = f'{path_to_save}/{file_name}.xlsx'
    session_name_index = 0
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        for sheet_name, df, in all_comparisons.items():
            df = df[['Driver 1', 'LapNumber D1', 'EvoCorrected D1', 'Compound', 'TyreLife',
                     'EvoCorrected D2', 'LapNumber D2', 'Driver 2', 'Exclude']]
            d1 = df['Driver 1'].loc[0]
            d2 = df['Driver 2'].loc[0]
            mean_values = df[df['Exclude'] != 'Yes'][['EvoCorrected D1', 'EvoCorrected D2']].mean()
            mean_row = pd.Series([None] * df.shape[1], index=df.columns)
            mean_row[['EvoCorrected D1', 'EvoCorrected D2']] = mean_values
            df = df._append(mean_row, ignore_index=True).sort_values(['LapNumber D1', 'LapNumber D2'], ascending=[True, True])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            green_fill = PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid')
            compound_color = {'MEDIUM': yellow_fill, 'HARD': white_fill, 'SOFT': red_fill, 'INTERMEDIATE': green_fill}
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if row[-1].value == 'Yes':
                        cell.fill = orange_fill
                    elif row[3].value != '':
                        cell.fill = compound_color[row[3].value]

            time_diff = mean_values[0] - mean_values[1]
            if time_diff < 0:
                message = f'{d1} FASTER BY'
            else:
                message = f'{d2} FASTER BY'

            print(f'{message} {abs(time_diff):.3f}s in {sessions_names[session_name_index]}')
            session_name_index += 1

            worksheet.cell(row=1, column=worksheet.max_column + 2, value=f"{d1} Average Time")
            worksheet.cell(row=2, column=worksheet.max_column, value=f"{d2} Average Time")
            worksheet.cell(row=1, column=worksheet.max_column + 1, value=f'{mean_values[0]:.3f}')
            worksheet.cell(row=2, column=worksheet.max_column, value=f'{mean_values[1]:.3f}')
            worksheet.cell(row=3, column=worksheet.max_column - 1, value=message)
            worksheet.cell(row=3, column=worksheet.max_column, value=f'{abs(time_diff):.3f}s')

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    cell.alignment = Alignment(horizontal='center')
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 5)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            for cell in worksheet[worksheet.max_row]:
                cell.font = openpyxl.styles.Font(bold=True)


        writer._save()

    if not all_teams:
        race_differences = []
        for l1, l2 in zip(total_laps_d1, total_laps_d2):
            delta_diff = (l1 - l2) / ((l1 + l2) / 2) * 100
            race_differences.append(delta_diff)

        median = statistics.median(race_differences)
        mean = statistics.mean(race_differences)
        trim_mean = stats.trim_mean(race_differences, 0.05)
        print(race_differences)
        print(f'MEDIAN DIFF {d1 if median < 0 else d2} FASTER: {median}')
        print(f'MEAN DIFF {d1 if mean < 0 else d2} FASTER: {mean}')
        print(f'TRIM MEAN DIFF {d1 if trim_mean < 0 else d2} FASTER: {trim_mean}')

def fuel_correct_factor(year, session_id=None):
    schedule = fastf1.get_event_schedule(year, include_testing=False)
    schedule = [1] if session_id is not None else schedule
    tyres = ['HARD', 'MEDIUM', 'SOFT']
    all_laps = pd.DataFrame()
    delta_dict = {}
    for i in range(len(schedule)):
        session = fastf1.get_session(year, i + 1 if session_id is None else session_id, 'R')
        session.load()
        session_fuel_correct_factor = []
        drivers = session.laps['Driver'].unique()
        for d in drivers:
            laps = session.laps.pick_driver(d).pick_wo_box().pick_quicklaps()
            laps = laps[laps['Compound'].isin(tyres)]
            stints = laps.groupby(['Stint', 'Compound'])['Compound'].value_counts().reset_index()
            for t in tyres:
                stint_tyre = stints[stints['Compound'] == t]
                if len(stint_tyre) > 1:
                    stint_data = laps.groupby(['Stint', 'Compound']).agg(Laps=('Compound', 'size')).reset_index()
                    stint_numbers = stint_data[stint_data['Compound'] == t].sort_values(by='Laps', ascending=False)['Stint'].values[0:2]
                    stint_1 = pd.DataFrame(laps[laps['Stint'] == stint_numbers[0]])[
                        ['Driver', 'TyreLife', 'LapNumber', 'LapTime']]
                    stint_2 = pd.DataFrame(laps[laps['Stint'] == stint_numbers[1]])[
                        ['Driver', 'TyreLife', 'LapNumber', 'LapTime']]
                    compare_stints = pd.merge(stint_1, stint_2, how='inner', on=['Driver', 'TyreLife'])

                    compare_stints['Evo Factor'] = ((compare_stints['LapTime_x'].apply(lambda x: x.total_seconds())
                                                   - compare_stints['LapTime_y'].apply(lambda x: x.total_seconds()))
                                                  / (compare_stints['LapNumber_y'] - compare_stints['LapNumber_x']))

                    compare_stints = compare_stints.rename(columns={'LapNumber_x': 'Lap First Stint',
                                                                    'LapNumber_y': 'Lap Second Stint',
                                                                    'LapTime_x': 'LapTime First Stint',
                                                                    'LapTime_y': 'LapTime Second Stint'
                                                                    })
                    compare_stints['LapTime First Stint'] = compare_stints['LapTime First Stint'].apply(lambda x: format_timedelta(x))
                    compare_stints['LapTime Second Stint'] = compare_stints['LapTime Second Stint'].apply(lambda x: format_timedelta(x))
                    all_laps = pd.concat([all_laps, compare_stints])

                    for value in compare_stints['Evo Factor']:
                        session_fuel_correct_factor.append(value)

        if len(session_fuel_correct_factor) > 25:
            print(f'TOTAL VALUES {len(session_fuel_correct_factor)}')
            truncated_mean = np.mean(session_fuel_correct_factor)
            with pd.ExcelWriter(f'../resources/evo_factor/{year}_{session_id}.xlsx', engine='openpyxl') as writer:
                all_laps.to_excel(writer, sheet_name='Evolution Factor', index=False)
                worksheet = writer.sheets['Evolution Factor']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        cell.alignment = Alignment(horizontal='center')
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 5)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                writer._save()
            print(truncated_mean)
            delta_dict[session_id] = truncated_mean
        else:
            delta_dict[session_id] = np.NaN

    df = pd.DataFrame([delta_dict]).T.reset_index(drop=True)
    df.columns = ['FuelFactor']
    df['Year'] = year
    df['Round'] = df.index + 1 if session_id is None else session_id
    existing_data = pd.read_csv('../resources/csv/Fuel_factor.csv')
    composite_pk_columns = ['Year', 'Round']
    merged = pd.merge(existing_data[composite_pk_columns], df, how='outer', on=composite_pk_columns, indicator=True)
    unique_rows = merged[merged['_merge'] == 'right_only'].drop(columns=['_merge']).reset_index(drop=True)
    unique_rows['ID'] = unique_rows.index + max(existing_data['ID']) + 1
    unique_rows = unique_rows[['ID', 'Year', 'Round', 'FuelFactor']]
    unique_rows.to_csv('../resources/csv/Fuel_factor.csv', mode='a', header=False, index=False)


def driver_fuel_corrected_laps(session, driver):
    laps = session.laps.pick_driver(driver).pick_quicklaps().pick_wo_box()[['Compound', 'LapNumber', 'LapTime']]
    fuel_factors = pd.read_csv('../resources/csv/Fuel_factor.csv')
    round_id = session.event.RoundNumber
    year = session.event.year
    race_fuel_factor = fuel_factors[(fuel_factors['Round'] == round_id) &
                                    (fuel_factors['Year'] == year)]['FuelFactor'][0]
    if pd.isna(race_fuel_factor):
        print('NO FUEL FACTOR')
        race_fuel_factor = 0.06

    laps['LapTime'] = laps['LapTime'].dt.total_seconds()
    laps['FuelCorrected'] = laps['LapTime'] - race_fuel_factor * (session.total_laps - laps['LapNumber'])
    # laps['FuelCorrected'] = pd.to_timedelta(laps['FuelCorrected'], unit='s')
    laps = laps.reset_index(drop=True)
    color_dict = {
        'HARD': 'white',
        'MEDIUM': 'yellow',
        'SOFT': 'red'
    }
    colors = [color_dict.get(i) for i in laps['Compound'].values]
    fig, ax = plt.subplots(figsize=(8, 8))
    for i in range(len(laps) - 1):
        plt.plot(laps['LapNumber'][i:i + 2], laps['FuelCorrected'][i:i + 2], marker='o', color=colors[i])

    def format_timedelta_to_mins(seconds, pos):
        mins = int(seconds // 60)
        secs = int(seconds % 60)
        millis = int((seconds - int(seconds)) * 1000)
        return f'{mins:01d}:{secs:02d}.{millis:03d}'

    formatter = FuncFormatter(format_timedelta_to_mins)
    plt.gca().yaxis.set_major_formatter(formatter)
    ax.set_xlabel("Lap", font='Fira Sans', fontsize=16)
    ax.set_ylabel("Fuel Corrected Lap Time", font='Fira Sans', fontsize=16)
    plt.grid(color='w', which='major', axis='x', linestyle='--', alpha=0.4)
    plt.grid(color='w', which='major', axis='y', linestyle='--', alpha=0.4)
    plt.xticks(font='Fira Sans', fontsize=14)
    plt.yticks(font='Fira Sans', fontsize=14)
    plt.title(f'{driver} FUEL CORRECTED LAPS IN {str(session.event.year) + " " + session.event.Country.upper() + " " + session.name.upper()}',
              font='Fira Sans', fontsize=20)
    sns.despine(left=True, bottom=True)
    plt.figtext(0.01, 0.02, '@F1BigData', font='Fira Sans', fontsize=17, color='gray', alpha=0.5)
    plt.tight_layout()
    plt.savefig(f'../PNGs/{driver} FUEL CORRECTED LAPS IN {str(session.event.year) + " " + session.event.Country + " " + session.name}.png', dpi=450)
    plt.show()


def race_diff_v2(year, round=None, save=False):
    schedule = fastf1.get_event_schedule(year, include_testing=False)
    schedule = [1] if round is not None else schedule
    tyres = ['SOFT', 'MEDIUM', 'HARD']
    fuel_factors = pd.read_csv('../resources/csv/Fuel_factor.csv')
    raw_race_pace = pd.read_csv('../resources/csv/Raw_race_pace_delta.csv')
    delta_by_race = {}
    sessions_names = []
    for i in range(len(schedule)):
        session = fastf1.get_session(year, round if round is not None else i + 1, 'R')
        session.load(messages=False)
        sessions_names.append(session.event['Location'].split('-')[0])
        race_fuel_factor = fuel_factors[(fuel_factors['Year'] == year) & (fuel_factors['Round'] == i + 1)]['FuelFactor']
        total_laps = pd.DataFrame(columns=['Team', 'LapTime', 'LapNumber', 'TyreLife', 'Driver'])
        if not pd.isna(race_fuel_factor.loc[0]):
            race_fuel_factor = race_fuel_factor.loc[0]
        else:
            race_fuel_factor = fuel_factors['FuelFactor'].median()
        laps = session.laps.pick_quicklaps().pick_wo_box()
        for tyre in tyres:
            laps_per_tyre = laps[laps['Compound'] == tyre][['Team', 'LapTime', 'LapNumber', 'TyreLife', 'Driver']]
            teams_per_tyre = laps_per_tyre['Team'].unique()
            if len(teams_per_tyre) < 10:
                print(f'NO DATA FOR {tyre} {session}')
            else:
                common_laps = laps_per_tyre.groupby('Team')['TyreLife'].apply(set)
                common_tyre_life = set.intersection(*common_laps)
                laps_per_tyre = laps_per_tyre[laps_per_tyre['TyreLife'].isin(common_tyre_life)]
                total_laps = pd.concat([laps_per_tyre, total_laps], ignore_index=True)
                print(f'{len(laps_per_tyre)} with {tyre} in {session}')
        if len(total_laps) > 225:
            total_laps['LapTime'] = total_laps['LapTime'].dt.total_seconds()
            total_laps['FuelCorrected'] = total_laps['LapTime'] - race_fuel_factor * (
                    session.total_laps - total_laps['LapNumber'])
            total_laps['FuelCorrected'] = pd.to_timedelta(total_laps['FuelCorrected'], unit='s')
            total_laps = total_laps.groupby(['Team', 'Driver'])['FuelCorrected'].mean().reset_index()
            total_laps = (total_laps.groupby('Team')['FuelCorrected'].min().reset_index()
                          .sort_values(by='FuelCorrected', ascending=True)).reset_index(drop=True)
            fastest_pace = total_laps['FuelCorrected'].loc[0]
            for t in session.laps['Team'].unique():
                try:
                    team_pace = total_laps[total_laps['Team'] == t]['FuelCorrected'].loc[0]
                    delta_diff = ((team_pace - fastest_pace) / fastest_pace) * 100
                except IndexError:
                    delta_diff = np.NaN
                if t not in delta_by_race:
                    delta_by_race[t] = [delta_diff]
                else:
                    delta_by_race[t].append(delta_diff)
        else:
            raw_data = raw_race_pace[(raw_race_pace['Year'] == year) & (raw_race_pace['Session'] == i + 1)]
            raw_teams = raw_data['Team'].unique()
            for t in raw_teams:
                team_pace = raw_data[raw_data['Team'] == t]['Delta'].iloc[0]
                if t not in delta_by_race:
                    delta_by_race[t] = [team_pace]
                else:
                    delta_by_race[t].append(team_pace)

    session_names = append_duplicate_number(sessions_names)
    print(delta_by_race)

    if save:
        df_save = pd.DataFrame(delta_by_race)
        df_save['Year'] = year
        df_save['Track'] = np.NaN
        df_save['Session'] = df_save.index + 1
        df_save = df_save.melt(id_vars=['Year', 'Session', 'Track'], var_name='Team', value_name='Delta')
        df_save = df_save[['Team', 'Delta', 'Year', 'Session', 'Track']]
        df_save.to_csv('../resources/csv/Intermediate_race_pace_delta.csv', index=False)

    fig, ax1 = plt.subplots(figsize=(12, 8))
    plt.rcParams["font.family"] = "Fira Sans"
    for team, deltas in delta_by_race.items():
        plt.plot(session_names, deltas,
                 label=team, marker='o', color=team_colors_2023.get(team), markersize=7, linewidth=3)
    plt.gca().invert_yaxis()
    plt.legend(loc='lower right', fontsize='medium')
    plt.title(f'{year} AVERAGE RACE PACE PER CIRCUIT', font='Fira Sans', fontsize=20)
    plt.ylabel('Percentage time difference (%)', font='Fira Sans', fontsize=16)
    plt.xlabel('Circuit', font='Fira Sans', fontsize=16)
    ax1.yaxis.grid(True, linestyle='--')
    ax1.xaxis.grid(True, linestyle='--', alpha=0.2)
    plt.xticks(rotation=90, fontsize=12, fontname='Fira Sans')
    plt.yticks(fontsize=12, fontname='Fira Sans')
    plt.tight_layout()
    plt.savefig(f"../PNGs/{year} race time difference.png", dpi=400)
    plt.show()


def percentage_race_ahead(start=2001, end=2100, year_drivers=None):
    ergast = My_Ergast()
    races = ergast.get_race_results([i for i in range(start, end)]).content
    circuits = ['catalunya']
    drivers_dict = {}
    for r in races:
        if len(r[r['circuitRef'].isin(circuits)]) > 0:
            drivers_in_q = r['fullName'].unique()
            for d in drivers_in_q:
                driver_data = r[r['fullName'] == d]
                driver_pos = driver_data['position'].loc[0]
                driver_status = driver_data[driver_data['position'] == driver_pos]['status'].loc[0]
                if ('Finished' in driver_status or '+' in driver_status) and len(driver_data) == 1:
                    driver_team = driver_data['constructorName'].loc[0]
                    team_data = r[r['constructorName'] == driver_team]
                    team_data = team_data[team_data['fullName'] != d]
                    for teammate in team_data['fullName'].unique():
                        teammate_pos = r[r['fullName'] == teammate]
                        if len(teammate_pos) == 1:
                            teammate_pos = teammate_pos['position'].loc[0]
                            teammate_status = r[(r['fullName'] == teammate)
                                                & (r['position'] == teammate_pos)]['status'].loc[0]
                            if 'Finished' in teammate_status or '+' in teammate_status:
                                win = 1
                                if driver_pos > teammate_pos:
                                    win = 0
                                if d not in drivers_dict:
                                    drivers_dict[d] = [win]
                                else:
                                    drivers_dict[d].append(win)
            print(f'{r["year"].loc[0]}: {r["circuitName"].loc[0]}')
    final_dict = {}
    h2h_dict = {}

    if year_drivers is not None:
        valid_drivers = []
        drivers = My_Ergast().get_qualy_results([year_drivers])
        for d in drivers.content:
            valid_drivers.extend(d['fullName'].values)

    for d, w in drivers_dict.items():
        percentage = round((sum(w) / len(w)) * 100, 2)
        if year_drivers is not None:
            if d in valid_drivers:
                final_dict[d] = percentage
                h2h_dict[d] = f'({sum(w)}/{len(w)})'
        else:
            final_dict[d] = percentage
            h2h_dict[d] = f'({sum(w)}/{len(w)})'

    final_dict = dict(sorted(final_dict.items(), key=lambda item: item[1], reverse=True))
    for d, w in final_dict.items():
        print(f'{d}: {w:.2f}% {h2h_dict[d]}')


def delta_reference_team(year, round=None):
    tyres = ['SOFT', 'MEDIUM', 'HARD']
    fuel_factors = pd.read_csv('../resources/csv/Fuel_factor.csv')
    final_delta = {}
    schedule = fastf1.get_event_schedule(year, include_testing=False)
    schedule = [1] if round is not None else schedule
    sessions_names = []
    for i in range(len(schedule)):
        delta_by_race = {}
        session = fastf1.get_session(year, i + 1 if round is None else round, 'R')
        session.load(messages=False)
        sessions_names.append(session.event['Location'].split('-')[0])
        round_number = session.event.RoundNumber
        teams = session.laps['Team'].unique()
        teams = [t for t in teams if not pd.isna(t)]
        race_fuel_factor = fuel_factors[(fuel_factors['Year'] == year) & (fuel_factors['Round'] == round_number)]['FuelFactor']
        total_laps = pd.DataFrame(columns=['Team', 'LapTime', 'LapNumber', 'TyreLife', 'Driver', 'Compound'])
        if not pd.isna(race_fuel_factor.loc[0]):
            race_fuel_factor = race_fuel_factor.loc[0]
        else:
            race_fuel_factor = fuel_factors['FuelFactor'].median()
        for t1 in teams:
            for t2 in teams:
                if t1 != t2:
                    for tyre in tyres:
                        laps = session.laps.pick_teams([t1, t2]).pick_quicklaps().pick_wo_box()
                        laps_per_tyre = laps[laps['Compound'] == tyre][
                            ['Team', 'LapTime', 'LapNumber', 'TyreLife', 'Driver', 'Compound']]
                        common_laps = laps_per_tyre.groupby('Team')['TyreLife'].apply(set)
                        if len(common_laps) > 0:
                            common_tyre_life = set.intersection(*common_laps)
                            laps_per_tyre = laps_per_tyre[laps_per_tyre['TyreLife'].isin(common_tyre_life)]
                            total_laps = pd.concat([laps_per_tyre, total_laps], ignore_index=True)

        total_laps = total_laps.drop_duplicates()
        max_tyre_age = pd.Series(total_laps.groupby('TyreLife').size())
        cut_value = np.mean(max_tyre_age)
        # cut_value = max(max_tyre_age[max_tyre_age >= cut_value].index.values)
        total_laps = total_laps[total_laps['TyreLife'].isin(max_tyre_age[max_tyre_age >= cut_value].index.values)]
        total_laps.loc[:, 'LapTime'] = total_laps['LapTime'].dt.total_seconds()
        total_laps.loc[:, 'FuelCorrected'] = total_laps['LapTime'] - race_fuel_factor * (
                session.total_laps - total_laps['LapNumber'])
        total_laps.loc[:, 'FuelCorrected'] = pd.to_timedelta(total_laps['FuelCorrected'], unit='s')
        total_laps = total_laps.groupby(['Team', 'Driver'])['FuelCorrected'].median().reset_index()
        teams_data = (total_laps.groupby('Team')['FuelCorrected'].min().reset_index()
                      .sort_values(by='FuelCorrected', ascending=True)).reset_index(drop=True)
        for t in teams:
            try:
                fastest_pace = teams_data['FuelCorrected'].loc[0]
                team_pace = teams_data[teams_data['Team'] == t]['FuelCorrected'].loc[0]
                delta_diff = ((team_pace - fastest_pace) / fastest_pace) * 100
            except IndexError:
                delta_diff = np.NaN

            delta_by_race[t] = delta_diff
        delta_by_race = dict(sorted(delta_by_race.items(), key=lambda x: x[1]))
        fastest_time = teams_data['FuelCorrected'].min()
        teams_data['Diff_in_seconds'] = teams_data['FuelCorrected'] - fastest_time
        teams_data['Diff_in_seconds'] = teams_data['Diff_in_seconds'].dt.total_seconds()
        fastest_time_seconds = fastest_time.total_seconds()
        teams_data['Percentage_diff'] = (teams_data['Diff_in_seconds'] / fastest_time_seconds) * 100
        for t, d, in delta_by_race.items():
            if t not in final_delta:
                final_delta[t] = [d]
            else:
                final_delta[t].append(d)

        print(teams_data)


    sessions_names = append_duplicate_number(sessions_names)
    fig, ax1 = plt.subplots(figsize=(12, 8))
    plt.rcParams["font.family"] = "Fira Sans"
    for team, deltas in final_delta.items():
        plt.plot(sessions_names, deltas,
                 label=team, marker='o', color=team_colors_2023.get(team), markersize=7, linewidth=3)
    plt.gca().invert_yaxis()
    plt.legend(loc='lower right', fontsize='medium')
    plt.title(f'{year} AVERAGE RACE PACE PER CIRCUIT', font='Fira Sans', fontsize=20)
    plt.ylabel('Percentage time difference (%)', font='Fira Sans', fontsize=16)
    plt.xlabel('Circuit', font='Fira Sans', fontsize=16)
    ax1.yaxis.grid(True, linestyle='--')
    ax1.xaxis.grid(True, linestyle='--', alpha=0.2)
    plt.xticks(rotation=90, fontsize=12, fontname='Fira Sans')
    plt.yticks(fontsize=12, fontname='Fira Sans')
    plt.tight_layout()
    plt.savefig(f"../PNGs/{year} race time difference.png", dpi=400)
    plt.show()

def avg_race_pos_dif(driver):

    races = My_Ergast().get_race_results([i for i in range(1950, 2100)])
    delta_per_year = {}
    for q in races.content:
        driver_data = q[q['fullName'] == driver]
        if len(driver_data) > 0:
            status = driver_data['status'].loc[0]
            if status == 'Finished' or '+' in status:
                driver_pos = driver_data['position'].loc[0]
                team = driver_data['constructorName'].loc[0]
                teammates = q[(q['constructorName'] == team) & (q['fullName'] != driver)]['fullName']
                for t in teammates.values:
                    teammate_data = q[q['fullName'] == t]
                    if len(teammate_data) == 1:
                        teammate_valid = teammate_data['status'].loc[0]
                        year = q['year'].loc[0]
                        if teammate_valid == 'Finished' or '+' in teammate_valid:
                            teammate_pos = teammate_data['position'].loc[0]
                            if (year, t) not in delta_per_year:
                                delta_per_year[(year, t)] = [[], []]
                            delta_per_year[(year, t)][0].append(driver_pos)
                            delta_per_year[(year, t)][1].append(teammate_pos)

    for y, d in delta_per_year.items():
        driver_avg = np.median(d[0])
        teammate_avg = np.median(d[1])
        diff = driver_avg - teammate_avg
        diff_str = f'+{diff:.2f}' if diff > 0 else f'{diff:.2f}'
        print(f'{"🔴" if diff > 0 else "🟢" if diff < 0 else "🟰"}{y[0]}: {y[1]} ({diff_str})')

    print(f'If the value is less than 0, {driver} has a better median race position.')
    print(f'If the value is greater than 0, {driver} has a worst median race position.')
    print('Only races in which BOTH drivers finished.')


def all_drivers_race_h2h(start=1950, end=2100):

    races = My_Ergast().get_race_results([i for i in range(start, end)])
    drivers_dict = {}
    for r in races.content:
        drivers = r['fullName'].unique()
        for d in drivers:
            driver_data = r[r['fullName'] == d]
            if len(driver_data) == 1:
                driver_status = driver_data['status'].loc[0]
                if driver_status == 'Finished' or '+' in driver_status:
                    driver_team = driver_data['constructorName'].loc[0]
                    teammates = r[(r['constructorName'] == driver_team) & (r['fullName'] != d)]['fullName'].unique()
                    for t in teammates:
                        teammate_data = r[r['fullName'] == t]
                        if len(teammate_data) == 1:
                            teammate_status = teammate_data['status'].loc[0]
                            if teammate_status == 'Finished' or '+' in teammate_status:
                                driver_pos = driver_data['position'].loc[0]
                                teammate_pos = teammate_data['position'].loc[0]
                                win_h2h = None
                                if driver_pos < teammate_pos:
                                    win_h2h = 1
                                elif driver_pos > teammate_pos:
                                    win_h2h = 0
                                if win_h2h is not None:
                                    if d not in drivers_dict:
                                        drivers_dict[d] = [win_h2h]
                                    else:
                                        drivers_dict[d].append(win_h2h)
    percentage_dict = {}
    for d, h in drivers_dict.items():
        diff = (np.sum(h)/len(h)) * 100
        percentage_dict[d] = diff

    df = pd.DataFrame(percentage_dict.items(), columns=['Driver', 'Percentage'])
    df['Rank'] = df['Percentage'].rank(method='min', ascending=False)
    df.sort_values(by='Rank', inplace=True)
    for index, row in df.iterrows():
        driver = row['Driver']
        percentage = row['Percentage']
        rank = int(row['Rank'])  # Convert to int for display purposes
        total_runs = sum(drivers_dict[driver])
        num_runs = len(drivers_dict[driver])
        print(f'{rank} - {driver}: {percentage:.2f}% ({total_runs}/{num_runs})')


def tyre_deg(session):

    fuel_factors = pd.read_csv('../resources/csv/Fuel_factor.csv')
    round_id = session.event.RoundNumber
    year = session.event.year
    total_laps = session.total_laps
    race_fuel_factor = fuel_factors[(fuel_factors['Round'] == round_id) &
                                    (fuel_factors['Year'] == year)]['FuelFactor'][0]

    deg_evo = pd.DataFrame(columns=['Driver', 'Compound', 'Stint', 'LapTime', 'EvoCorrected', 'Deg_factor'])

    if pd.isna(race_fuel_factor):
        race_fuel_factor = 0.06
    print(race_fuel_factor)
    drivers = set(session.laps['Driver'])

    def remove_outliers_by_IQR(data, column_name):
        Q1 = data[column_name].quantile(0.25)
        Q3 = data[column_name].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR
        return data[(data[column_name] >= lower_bound) & (data[column_name] <= upper_bound)]

    for d in drivers:
        d_laps = session.laps.pick_driver(d).pick_wo_box().pick_quicklaps()
        d_laps['EvoCorrected'] = (d_laps['LapTime'].dt.total_seconds() - race_fuel_factor *
                                      (total_laps - d_laps['LapNumber']))

        stints = set(d_laps['Stint'].values)
        for s in stints:
            stint_laps = d_laps[d_laps['Stint'] == s].sort_values(by='LapNumber', ascending=True)
            df_to_append = pd.DataFrame(stint_laps[['Driver', 'Compound', 'Stint', 'LapTime', 'EvoCorrected']])
            stint_laps_clean = remove_outliers_by_IQR(df_to_append, 'EvoCorrected')
            if len(stint_laps_clean) >= 10:
                X = sm.add_constant(np.arange(len(df_to_append)))  # Use lap index as predictor
                y = df_to_append['EvoCorrected']
                huber = RLM(y, X, M=sm.robust.norms.HuberT()).fit()
                df_to_append.loc[df_to_append.index.isin(stint_laps_clean.index), 'Deg_factor'] = huber.params[1]
                deg_evo = deg_evo._append(df_to_append)

    deg_evo = deg_evo.groupby(['Compound', 'Driver', 'Stint'])['Deg_factor'].mean().reset_index()
    deg_evo['Driver_Stint'] = deg_evo['Driver'] + ' (' + deg_evo['Stint'].astype(int).astype(str) + ')'
    deg_evo = deg_evo.sort_values(by='Deg_factor', ascending=True)

    for tyre in ['HARD', 'MEDIUM', 'SOFT']:
        df_to_plot = deg_evo[deg_evo['Compound'] == tyre]
        if len(df_to_plot) > 0:
            fix, ax = plt.subplots(figsize=(9, 9))
            bars = plt.bar(df_to_plot['Driver_Stint'], df_to_plot['Deg_factor'])
            colors = [driver_colors.get(year).get(i) for i in df_to_plot['Driver']]
            round_bars(bars, ax, colors, color_1=None, color_2=None, y_offset_rounded=0.03, corner_radius=0.050, linewidth=4)
            annotate_bars(bars, ax, 0.002, 11, text_annotate='default', ceil_values=False, round=2,
                          y_negative_offset=-0.01, annotate_zero=False, negative_offset=0)

            plt.title(f'TYRE DEG WITH THE {tyre.upper()} COMPOUND', font='Fira Sans', fontsize=24)
            plt.ylabel('Average time loss per lap (s)', font='Fira Sans', fontsize=20)
            plt.xticks(rotation=90, font='Fira Sans', fontsize=18)
            plt.yticks(font='Fira Sans', fontsize=16)
            plt.ylim(bottom=min(df_to_plot['Deg_factor']) - 0.05, top=max(df_to_plot['Deg_factor']) + 0.05)
            color_index = 0
            for label in ax.get_xticklabels():
                label_text = label.get_text()
                if len(label_text) > 3:
                    label.set_text(label_text[:3])
                label.set_color('white')
                label.set_fontsize(16)
                for_color = colors[color_index]
                if for_color == '#ffffff':
                    for_color = '#FF7C7C'
                label.set_path_effects([path_effects.withStroke(linewidth=2, foreground=for_color)])
                color_index += 1
            ax.yaxis.grid(True, linestyle='--', alpha=0.5)
            plt.tight_layout()
            plt.savefig(f'../PNGs/TYRE DEG WITH THE {tyre.upper()} COMPOUND.png', dpi=450)
            plt.show()

